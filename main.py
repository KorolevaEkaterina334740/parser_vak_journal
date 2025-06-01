import os
import random
import time
import traceback

import pandas as pd
from dotenv import load_dotenv
from fake_useragent import UserAgent
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait

load_dotenv()

ELIBRARY_LOGIN = os.getenv("ELIBRARY_LOGIN")
ELIBRARY_PASSWORD = os.getenv("ELIBRARY_PASSWORD")

LOGIN_URL = "https://elibrary.ru/defaultx.asp"
JOURNALS_URL = "https://www.elibrary.ru/titles.asp"
JOURNAL_PROFILE_URL = "https://www.elibrary.ru/title_profile.asp?id="
JOURNAL_ARTICLES_URL = "https://www.elibrary.ru/title_items.asp?id="
ARTICLE_BASE_URL = "https://www.elibrary.ru/item.asp?id="

INDEX_VAK = 3
JOURNALS_CATEGORY = 'Мультидисциплинарные журналы по всем направлениям науки' + '  ' + '(1627)'


def get_driver():
    user_agent = UserAgent().random
    options = webdriver.ChromeOptions()
    options.add_argument(f"user-agent={user_agent}")
    return webdriver.Chrome(options=options)


def login(driver):
    driver.get(LOGIN_URL)
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.NAME, "login")))
    driver.find_element(By.ID, "login").send_keys(ELIBRARY_LOGIN)
    driver.find_element(By.ID, "password").send_keys(ELIBRARY_PASSWORD)
    driver.find_element(By.XPATH, "//div[@class='butred' and contains(text(), 'Вход')]").click()
    print("Успешная авторизация")


def select_filters(driver):
    driver.get(JOURNALS_URL)
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.NAME, "vak")))

    Select(driver.find_element(By.NAME, "rubriccode")).select_by_visible_text(JOURNALS_CATEGORY)
    # .select_by_index(63))
    Select(driver.find_element(By.NAME, "vak")).select_by_index(1 + INDEX_VAK)
    driver.find_element(By.CSS_SELECTOR, "div.butred[onclick='title_search()']").click()


def parse_journals_table(driver):
    journals = []
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "restab")))
    time.sleep(random.uniform(5, 7))

    rows = driver.find_element(By.ID, "restab").find_elements(By.TAG_NAME, "tr")[3:60]
    for row in rows:
        cells = row.find_elements(By.TAG_NAME, "td")
        jid = row.get_attribute("id")[1:]
        journals.append({
            "id": jid,
            "link": f"{JOURNAL_PROFILE_URL}{jid}",
            "article_link": f"{JOURNAL_ARTICLES_URL}{jid}",
            "title": cells[2].text.split('\n')[0],
            "author": cells[2].text.split('\n')[1],
            "publications": cells[3].text,
            "article": cells[4].text,
            "quotes": cells[5].text
        })
    return journals


def parse_journal_detail(driver, journal):
    data = {}
    driver.get(journal["link"])
    time.sleep(random.uniform(15, 30))
    try:
        table = WebDriverWait(driver, 10).until(EC.presence_of_element_located(
            (By.XPATH, "(//table[@width='580' and @cellspacing='0' and @cellpadding='3'])[2]")))
        rows = table.find_elements(By.TAG_NAME, "tr")

        def safe_get(row, idx):
            try:
                return row.find_elements(By.TAG_NAME, "td")[idx].text
            except:
                return ""

        count_articles = safe_get(rows[3], -1)
        views_year = safe_get(rows[59], -1)

        try:
            views_per_article = int(views_year) / int(count_articles)
        except:
            views_per_article = 0

        data.update({
            'link': journal["link"],
            'category': JOURNALS_CATEGORY,
            'title': journal["title"],
            'author': journal["author"],
            'vak': INDEX_VAK,
            'publications': journal["publications"],
            'article': journal["article"],
            'quotes': journal["quotes"],
            'science_index': safe_get(rows[5], -1),
            'index_hirsha': safe_get(rows[46], -1),
            'index_herfindal': safe_get(rows[49], -1),
            'index_jinny': safe_get(rows[53], -1),
            'views_per_year': views_year,
            'count_of_articles': count_articles,
            'views_per_article': views_per_article
        })
        return data
    except TimeoutException:
        print(f"Timeout при загрузке данных журнала {journal['title']}")
        return None


def parse_articles(driver, article_link, journal_title):
    articles = []
    driver.get(article_link)
    time.sleep(random.uniform(15, 30))
    try:
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "hdr_years"))).click()
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "year_2023"))).click()

        time.sleep(random.uniform(3, 7))
        driver.find_element(By.CSS_SELECTOR, "div.butred[onclick='pub_search()']").click()

        time.sleep(random.uniform(3, 7))
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, 'restab')))

        list_articles = driver.find_element(By.ID, 'restab').find_elements(By.TAG_NAME, 'tr')[3:23]
        rows = [item.get_attribute("id")[3:] for item in list_articles]

        for aid in rows:
            link = f"{ARTICLE_BASE_URL}{aid}"

            driver.get(link)
            time.sleep(random.uniform(15, 30))

            try:
                title_table = driver.find_element(
                    By.XPATH,
                    "(//table[@width='550' and @cellspacing='0' and @cellpadding='3' and @border='0'])[2]"
                )
                title = title_table.find_element(By.TAG_NAME, "tr").find_elements(By.TAG_NAME, "td")[1].text

                author_table = driver.find_element(
                    By.XPATH,
                    "(//table[@width='550' and @cellspacing='0' and @cellpadding='3' and @border='0'])[3]"
                )
                author = author_table.find_element(By.TAG_NAME, "tr").find_elements(By.TAG_NAME, "td")[1].text.split('\n')[0]

                keywords_table = driver.find_element(
                    By.XPATH,
                    "(//table[@width='550' and @border='0' and @cellspacing='0' and @cellpadding='3'])[5]"
                )
                keywords = keywords_table.find_elements(By.TAG_NAME, "tr")[1].find_elements(By.TAG_NAME, "td")[1].text

                category_table = driver.find_element(
                    By.XPATH,
                    "(//table[@width='100%' and @border='0' and @cellspacing='0' and @cellpadding='3'])[2]"
                )

                category = category_table.find_element(By.TAG_NAME, "tr").find_elements(By.TAG_NAME, "td")[1].text

                library_table = driver.find_element(
                    By.XPATH,
                    "(//table[@width='550' and @border='0' and @cellspacing='0' and @cellpadding='3'])[6]"
                ).find_element(
                    By.XPATH,
                    "(//table[@width='100%' and @border='0' and @cellspacing='0' and @cellpadding='3'])"
                )
                library_tr_rows = library_table.find_elements(By.TAG_NAME, "tr")

                include_RINC = library_tr_rows[0].find_elements(By.TAG_NAME, "td")[0].text.split(':')[1].strip(),
                quotas_in_RINC = library_tr_rows[0].find_elements(By.TAG_NAME, "td")[1].text.split(':')[1].strip(),
                include_core_RINC = library_tr_rows[1].find_elements(By.TAG_NAME, "td")[0].text.split(':')[1].strip(),
                qoutas_core_RINC = library_tr_rows[1].find_elements(By.TAG_NAME, "td")[1].text.split(':')[1].strip(),
                reviews = library_tr_rows[2].find_elements(By.TAG_NAME, "td")[0].text.split(':')[1].strip(),
                percent_in_top_SI = library_tr_rows[2].find_elements(By.TAG_NAME, "td")[1].text.split(':')[1].strip(),

                stats_table = driver.find_element(
                    By.XPATH,
                    "(//table[@width='100%' and @border='0' and @cellspacing='0' and @cellpadding='3'])[7]"
                )
                stats = stats_table.find_elements(By.TAG_NAME, "tr")

                articles.append({
                    'link': link,
                    "title": title,
                    "author": author,
                    "keywords": keywords,
                    "category": category,

                    'include_RINC': include_RINC[0],
                    'quotas_in_RINC': quotas_in_RINC[0],
                    'include_core_RINC': include_core_RINC[0],
                    'qoutas_core_RINC': qoutas_core_RINC[0],
                    'reviews': reviews[0],
                    'percent_in_top_SI': percent_in_top_SI[0],

                    "views": stats[0].find_elements(By.TAG_NAME, "td")[0].text.split(':')[1].split()[0].strip(),
                    "downloads": stats[0].find_elements(By.TAG_NAME, "td")[1].text.split(':')[1].split()[0].strip(),
                    "collections": stats[0].find_elements(By.TAG_NAME, "td")[2].text.split(':')[1].strip(),
                    "total_score": stats[1].find_elements(By.TAG_NAME, "td")[0].text.split(':')[1].strip(),
                    "avg_score": stats[1].find_elements(By.TAG_NAME, "td")[1].text.split(':')[1],
                    "comments": stats[1].find_elements(By.TAG_NAME, "td")[2].text.split(':')[1].strip(),
                })
                time.sleep(random.uniform(15, 30))

            except Exception as e:
                print(f"Ошибка при обработке статьи {link}: {e}")

        if articles:
            df = pd.DataFrame(articles)
            with pd.ExcelWriter("journals_data.xlsx", engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                sheet_name = journal_title[:30].replace(" ", "_")
                if sheet_name in writer.book.sheetnames:
                    sheet = writer.book[sheet_name]
                    startrow = sheet.max_row
                else:
                    startrow = 0

                df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=startrow)

            wb = load_workbook("journals_data.xlsx")
            ws = wb[journal_title[:30].replace(" ", "_")]

            header_font = Font(bold=True)
            fill = PatternFill("solid", fgColor="D7E4BC")
            align = Alignment(horizontal="center", vertical="top", wrap_text=True)
            border = Border(bottom=Side(border_style="thin", color="000000"))

            for col_num, column_title in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = fill
                cell.alignment = align
                cell.border = border

                max_length = max(
                    [len(str(cell.value)) for cell in ws[get_column_letter(col_num)] if cell.value is not None] + [
                        len(column_title)]
                )
                ws.column_dimensions[get_column_letter(col_num)].width = max_length + 2

            wb.save("journals_data.xlsx")
            print(f"Сохранено {len(articles)} статей для журнала {journal_title}")

    except Exception as e:
        print(f"Ошибка при сборе статей для журнала {journal_title}: {e}")


def save_to_excel(records, path):
    df = pd.DataFrame(records)
    with pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        sheet_name = 'Журналы'
        if sheet_name in writer.book.sheetnames:
            sheet = writer.book[sheet_name]
            startrow = sheet.max_row
        else:
            startrow = 0
        df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=startrow)

    wb = load_workbook(path)
    ws = wb['Журналы']

    header_font = Font(bold=True)
    fill = PatternFill("solid", fgColor="D7E4BC")
    align = Alignment(horizontal="center", vertical="top", wrap_text=True)
    border = Border(bottom=Side(border_style="thin", color="000000"))

    for col_num, column_title in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = align
        cell.border = border

        max_length = max(
            [len(str(cell.value)) for cell in ws[get_column_letter(col_num)] if cell.value is not None] + [
                len(column_title)]
        )
        ws.column_dimensions[get_column_letter(col_num)].width = max_length + 2

    wb.save(path)
    print(f"Файл успешно сохранён: {path}")


def main():
    driver = get_driver()
    records = []
    try:
        time.sleep(random.uniform(10, 20))
        login(driver)

        time.sleep(random.uniform(10, 20))
        select_filters(driver)

        time.sleep(random.uniform(10, 20))
        journals = parse_journals_table(driver)

        for journal in journals:
            time.sleep(random.uniform(15, 30))
            try:
                record = parse_journal_detail(driver, journal)
                if record:
                    records.append(record)

                    #time.sleep(random.uniform(10, 15))
                    #parse_articles(driver, journal["article_link"], journal["title"])
            except Exception as e:
                print(f"Ошибка при обработке журнала: {journal['title']}")
                print(traceback.format_exc())
    except Exception as e:
        print(f"Общая ошибка: {e}")
        print(traceback.format_exc())
    finally:
        driver.quit()
        if records:
            save_to_excel(records, "journals_data.xlsx")
        else:
            print("Данных для сохранения нет")


if __name__ == '__main__':
    main()
